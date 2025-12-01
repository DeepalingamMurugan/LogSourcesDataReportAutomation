import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwings as xw
import streamlit as st
import tempfile
from datetime import datetime


def qradintfun():

    
    st.title("QRadar Integration Report Processor")
    
    st.write("Build for auto filtration and modification of data, works in tempfile each write will be temporary. download the file at end the process. do not interrupt.")

    # Streamlit file uploader
    file_asset = st.file_uploader("Upload asset list xlsx", type=['xlsx'])
    file_log = st.file_uploader("Upload log source xlsx", type=['xlsx'])
    file_qformula = st.file_uploader("Upload formula sheet xlsx", type=['xlsx'])

    # put imported files to temporary path
    if file_asset and file_log and file_qformula and st.button("Run process"):

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_asset_file:
            temp_asset_file.write(file_asset.read())
            temp_asset_path = temp_asset_file.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_log_file:
            temp_log_file.write(file_log.read())
            temp_log_path = temp_log_file.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_qformula_file:
            temp_qformula_file.write(file_qformula.read())
            temp_qformula_path = temp_qformula_file.name


        #================================= assetList&TelephonyFilteration =================================================

        df = pd.read_excel(temp_asset_path, keep_default_na=False)

        # Remove the 'artifact' column
        if 'artifact' in df.columns:
            df = df.drop(columns=['artifact'])

        # Filter the 'Status' column for 'decommissioned' and 'Inactive' and remove corresponding rows
        status_to_remove = ['Decommissioned', 'Inactive']
        df = df[~df['Status'].isin(status_to_remove)]

        # Filter the 'Owner' column for 'Sogeti SOC' and remove corresponding rows
        owner_to_remove = 'Sogeti SOC'
        df = df[df['Owner'] != owner_to_remove]

        # Remove extra spaces from 'IP Address' and 'Hostname' columns
        df['IP Address'] = df['IP Address'].str.strip()
        df['Hostname'] = df['Hostname'].str.strip()

        # Filter 'Device Type' for 'Telephony' and create a separate DataFrame for 'Telephony' devices
        telephony_df = df[df['Device Type'] == 'Telephony']

        # Apply all the previous operations to the 'Telephony' DataFrame
        telephony_df = telephony_df[~telephony_df['Status'].isin(status_to_remove)]
        telephony_df = telephony_df[telephony_df['Owner'] != owner_to_remove]
        telephony_df['IP Address'] = telephony_df['IP Address'].str.strip()
        telephony_df['Hostname'] = telephony_df['Hostname'].str.strip()

        # Remove 'Telephony' devices from the original DataFrame
        df = df[df['Device Type'] != 'Telephony']

        # Save the modified original DataFrame back to the 'asset.xlsx' file
        with pd.ExcelWriter(temp_asset_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Asset List', index=False)

        # Save the 'Telephony' DataFrame to a new sheet in the same Excel file
        with pd.ExcelWriter(temp_asset_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            telephony_df.to_excel(writer, sheet_name='Telephony', index=False)

        st.info("Dropped artifact, inactive/decommissioned status, sogetic SOC. stripped IP addr & hostname. Applied for Asset List and Telephony worksheets.")

        #st.success("Data process successful for Asset List & ready for download.")

        #with open(temp_asset_path, 'rb') as f:
        #    st.download_button('Download processed Asset List xlsx for reference', f, file_name="processed_asset.xlsx")

        #st.info("Processing....")

        #==============================================================================================

        #===================================== assetListToQformulaCopy =================================

        # Read the 'Asset List' spreadsheet
        asset_df = pd.read_excel(temp_asset_path, keep_default_na=False, sheet_name='Asset List')

        # Open the 'qformula.xlsx' file with openpyxl
        qformula_wb = openpyxl.load_workbook(temp_qformula_path)

        # Get the 'Asset List' sheet from 'qformula.xlsx'
        qformula_ws = qformula_wb['Asset List']

        # Get the column titles from both dataframes
        asset_titles = asset_df.columns.tolist()
        qformula_titles = [cell.value for cell in qformula_ws[1] if cell.value is not None]

        # Find the matching columns, ignoring case
        matching_columns = [col for col in asset_titles if col.lower() in [title.lower() for title in qformula_titles]]

        # Copy and paste the matching columns row by row
        for col in matching_columns:
            col_index_asset = asset_titles.index(col)
            col_index_qformula = qformula_titles.index(col)

            # Copy data from 'asset.xlsx' and paste it into 'qformula.xlsx' row by row
            for row_index, row_data in enumerate(dataframe_to_rows(asset_df[[col]], index=False, header=False), start=2):
                for index, value in enumerate(row_data, start=1):
                    qformula_ws.cell(row=row_index, column=col_index_qformula + index, value=value)

        # Save the modified 'qformula.xlsx' file
        qformula_wb.save(temp_qformula_path)

        # Close the 'qformula.xlsx' file
        qformula_wb.close()

        st.info("Data copied from Asset List worksheet to Formula sheet. Temporary save!")

        #==============================================================================================

        #===================================== telephonyToQformulaCopy ==================================

        # Read the 'Asset List' spreadsheet from 'asset.xlsx'
        asset_df = pd.read_excel(temp_asset_path, keep_default_na=False, sheet_name='Telephony')

        # Open the 'qformula.xlsx' file with openpyxl
        qformula_wb = openpyxl.load_workbook(temp_qformula_path)

        # Get the 'Asset List' sheet from 'qformula.xlsx'
        qformula_ws = qformula_wb['Telephony']

        # Get the column titles from both dataframes
        asset_titles = asset_df.columns.tolist()
        qformula_titles = [cell.value for cell in qformula_ws[1] if cell.value is not None]

        # Find the matching columns, ignoring case
        matching_columns = [col for col in asset_titles if col.lower() in [title.lower() for title in qformula_titles]]

        # Copy and paste the matching columns row by row
        for col in matching_columns:
            col_index_asset = asset_titles.index(col)
            col_index_qformula = qformula_titles.index(col)

            # Copy data from 'asset.xlsx' and paste it into 'qformula.xlsx' row by row
            for row_index, row_data in enumerate(dataframe_to_rows(asset_df[[col]], index=False, header=False), start=2):
                for index, value in enumerate(row_data, start=1):
                    qformula_ws.cell(row=row_index, column=col_index_qformula + index, value=value)

        # Save the modified 'qformula.xlsx' file
        qformula_wb.save(temp_qformula_path)

        # Close the 'qformula.xlsx' file
        qformula_wb.close()

        st.info("Data copied from Telephony worksheet to Formula sheet. Temporary save!")

        #==============================================================================================
        
        #===================================== logsourceDateFormatting ===============================
        
        # Read the log file
        df_log = pd.read_excel(temp_log_path, keep_default_na=False)

        # Check if the first row is the header
        #headers = df_log.columns.tolist()
        #st.write("Headers:", headers)

        def format_date(date_str):
            if isinstance(date_str, str):
                date_parts = date_str.split()
                cleaned_date_str = ' '.join(date_parts[:3])
                return cleaned_date_str

        df_log['Last Event'] = df_log['Last Event'].apply(format_date)

        def change_date_format(date_str):
            try:
                if date_str:
                    formatted_date = pd.to_datetime(date_str, format='%b %d, %Y').strftime('%d-%b-%y')
                    return formatted_date
                else:
                    return None
            except ValueError:
                return "didnt work, value error"

        df_log['Last Event'] = df_log['Last Event'].apply(change_date_format)

        # Save the modified DataFrame back to the existing file
        with pd.ExcelWriter(temp_log_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_log.to_excel(writer, sheet_name='log-sources', index=False)

        st.info("Formatted date on log source. Temporary save!")

        #st.write(df_log.head())

        #=============================================================================================

        #===================================== logsourceFiltration ===================================

        # Read the 'log-sources'
        dq = pd.read_excel(temp_log_path, sheet_name='log-sources', keep_default_na=False)

        # Define the set of characters or values to remove
        values_to_remove_from_logSourceIdentifier = ['/* list rm from ide */']

        # Remove the values from the 'Log Source Identifier' column
        for value in values_to_remove_from_logSourceIdentifier:
            dq['Log Source Identifier'] = dq['Log Source Identifier'].str.replace(value, '', regex=False)

        # Save the modified DataFrame back to the Excel file
        with pd.ExcelWriter(temp_log_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            dq.to_excel(writer, sheet_name='log-sources', index=False)

        st.info("Strings replaced for specified identifiers & Applied to log source sheet.")

        #st.success("Data process successful for Log source & ready for download.")
        
        #with open(temp_log_path, 'rb') as f:
        #    st.download_button('Download processed Log source xlsx for reference', f, file_name="processed_logsource.xlsx")

        #st.info("Processing....")

        #==============================================================================================

        #===================================== logsourceToQformulaCopy =================================

        # Read the 'log-sources' spreadsheet 
        log_df = pd.read_excel(temp_log_path, keep_default_na=False, sheet_name='log-sources')

        # Open the 'qformula.xlsx' file with openpyxl
        qformula_wb = openpyxl.load_workbook(temp_qformula_path)

        # Get the 'log-sources' sheet from 'qformula.xlsx'
        qformula_ws = qformula_wb['Log Source']

        # Get the column titles from both dataframes
        log_titles = log_df.columns.tolist()
        qformula_titles = [cell.value for cell in qformula_ws[1] if cell.value is not None]

        # Find the matching columns, ignoring case
        matching_columns = [col for col in log_titles if col.lower() in [title.lower() for title in qformula_titles]]

        # Copy and paste the matching columns row by row
        for col in matching_columns:
            col_index_log = log_titles.index(col)
            col_index_qformula = qformula_titles.index(col)

            # Copy data from 'asset.xlsx' and paste it into 'qformula.xlsx' row by row
            for row_index, row_data in enumerate(dataframe_to_rows(log_df[[col]], index=False, header=False), start=2):
                for index, value in enumerate(row_data, start=1):
                    qformula_ws.cell(row=row_index, column=col_index_qformula + index, value=value)

        # Save the modified 'qformula.xlsx' file
        qformula_wb.save(temp_qformula_path)

        # Close the 'qformula.xlsx' file
        qformula_wb.close()

        st.info("Data copied from Log source sheet to Formula sheet. Temporary save!")

        #==================================================================================================

        # ==================================Qformula_formulaUpdationNA======================================

        # Load the workbook with openpyxl
        wb = openpyxl.load_workbook(temp_qformula_path, data_only=False)
        sheet = wb['Asset List']  # Adjust sheet name if necessary

        # Open the workbook with xlwings to evaluate formulas
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)
        sheet_xlwings = wb_xlwings.sheets['Asset List']

        # Iterate over the rows and check column G (Last event) for formula results
        for row_idx in range(2, sheet.max_row + 1):  # Assuming row 1 is the header
            col_g_cell = f'G{row_idx}'

            # Evaluate the formula in column G
            col_g_value = sheet_xlwings.range(col_g_cell).formula
            #print(col_g_value)
            col_g_result = sheet_xlwings.range(col_g_cell).value  # Evaluate the result of the formula
            #print(col_g_result)
            if col_g_result == "NA":  # Check if the result is 'NA'
                col_b_cell = f'B{row_idx}'  # Column B is the 2nd column
                col_f_cell = f'F{row_idx}'  # Column F is the 6th column

                # Update the formula in column A (Concat) to B2&F2
                concat_formula = f'={col_b_cell}&{col_f_cell}'
                sheet_xlwings.range(f'A{row_idx}').formula = concat_formula
        # Save and close the updated workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        st.info("Formula updation with hostname on NA events checks after IP addr. Temporary save!")

        #====================================================================================================

        #======================================Qformula_formulaUpdationNA2vslookup===========================
        # Load the workbook with openpyxl
        wb = openpyxl.load_workbook(temp_qformula_path, data_only=False)
        sheet = wb['Asset List']
        
        # Open the workbook with xlwings to evaluate formulas
        app = xw.App(visible=False)
        wb = app.books.open(temp_qformula_path)
        xlwing_sheet = wb.sheets["Asset List"]  # The name of your sheet
        
        # Iterate through column G ("Last event") and check for "NA" result
        for row in range(2, sheet.max_row + 1):  # Row indices start at 2
            cell = f'G{row}'
            # Get the formula as a string
            cell_value = xlwing_sheet.range(cell).formula

            # Evaluate the formula result
            eval_value = xlwing_sheet.range(cell).value

            # Check if the evaluated result is "NA"
            if eval_value == "NA":
                # Apply the new formula based on E and F values for the row
                new_formula = f'=IFERROR(VLOOKUP(E{row}, \'Log Source\'!D:G, 4, FALSE), IFERROR(VLOOKUP(F{row}, \'Log Source\'!D:G, 4, FALSE), "NA"))'
                xlwing_sheet.range(cell).formula = new_formula
                #print(f"Updated formula in Row {row}")

        # Save and close the workbook
        wb.save()
        wb.close()
        app.quit()
        
        st.info("Formula updation direct check with Log source on NA events checks after IP addr & hostname. Temporary save!")

        #===================================================================================================
        
        #=========================================formula_to_result=========================================

        # Open the workbook with xlwings
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)
        sheet_xlwings = wb_xlwings.sheets['Asset List']

        # Iterate over the rows and process based on conditions
        for row_idx in range(2, sheet_xlwings.range('A1').end('down').row + 1):  # Dynamically find last row
            col_g_cell = f'G{row_idx}'
            col_l_cell = f'L{row_idx}'

            # Evaluate the formula and get the result
            col_g_result = sheet_xlwings.range(col_g_cell).value
            col_l_value = sheet_xlwings.range(col_l_cell).value

            # Define cells to update
            cells_to_update = ['A', 'B', 'G', 'H', 'I']

            # Check conditions and update cells accordingly
            if col_g_result != "NA" or col_l_value in ["AVI Controller", "AVI Service Engine(LB)", "AVI Service Engine(RP)", "Firewall", "IPS/IDS"]:
                for col in cells_to_update:
                    cell_address = f'{col}{row_idx}'
                    # Get the result and replace formula with result
                    cell_result = sheet_xlwings.range(cell_address).value
                    sheet_xlwings.range(cell_address).value = cell_result

        # Save and close the updated workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        st.info("Replaced data with original value results based on condition if only values != NA or Net Devices. Temporary save!")

        #====================================================================================================

        #=======================================netDevStatusChange===========================================
        # Open the workbook with xlwings
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)
        sheet_xlwings = wb_xlwings.sheets['Asset List']

        # Iterate over the rows and process based on conditions
        for row_idx in range(2, sheet_xlwings.range('A1').end('down').row + 1):  # Dynamically find last row
            col_l_cell = f'L{row_idx}'

            # Read the value from column L
            col_l_value = sheet_xlwings.range(col_l_cell).value

            # Check if column L value is in the specified list
            if col_l_value in ["AVI Controller", "AVI Service Engine(LB)", "AVI Service Engine(RP)", "Firewall", "IPS/IDS"]:
                # Update column H and I with the specified values
                sheet_xlwings.range(f'H{row_idx}').value = "Fully Integrated"
                sheet_xlwings.range(f'I{row_idx}').value = "Sending logs"

        # Save and close the updated workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        st.info("Updated 'Integration Status' and 'Log Sending Status' on 'AVI, FW, IPS/IDS' - device type even if NA. Temporary save!")

        #====================================================================================================

        #============================droppingUnwantedRows====================================================

        # Open the workbook with xlwings
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)
        sheet_xlwings = wb_xlwings.sheets['Asset List']  # Adjust sheet name if necessary


        # Drop the columns "Concat", "Device OS", "Asset Integrated in QRadar", "Logs Sending Status"
        columns_to_delete = ['A', 'A', 'J', 'J']
        for col in reversed(columns_to_delete):
            sheet_xlwings.range(f'{col}:{col}').delete()

        # Save and close the updated workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        st.info("Dropped unwanted data from Formula sheet. Temporary save!")

        #===================================================================================================
        
        #===================================TelephonyNAchecktoLS============================================
        # Load the workbook with openpyxl
        wb = openpyxl.load_workbook(temp_qformula_path, data_only=False)
        sheet = wb['Telephony']
        
        # Open the workbook with xlwings to evaluate formulas
        app = xw.App(visible=False)
        wb = app.books.open(temp_qformula_path)
        xlwing_sheet = wb.sheets["Telephony"]  # The name of your sheet
        
        # Iterate through column G ("Last event") and check for "NA" result
        for row in range(2, sheet.max_row + 1):  # Row indices start at 2
            cell = f'E{row}'
            # Get the formula as a string
            cell_value = xlwing_sheet.range(cell).formula

            # Evaluate the formula result
            eval_value = xlwing_sheet.range(cell).value

            # Check if the evaluated result is "NA"
            if eval_value == "NA":
                # Apply the new formula based on E and F values for the row
                new_formula = f'=IFERROR(VLOOKUP(C{row}, \'Log Source\'!D:G, 4, FALSE), IFERROR(VLOOKUP(D{row}, \'Log Source\'!D:G, 4, FALSE), "NA"))'
                xlwing_sheet.range(cell).formula = new_formula
                #print(f"Updated formula in Row {row}")

        # Save and close the workbook
        wb.save()
        wb.close()
        app.quit()
        
        st.info("Formula updation direct check with Log source on NA events on Telephony worksheet. Temporary save!")
        
        #===================================================================================================

        #=====================================formula_to_result============================================

        # Open the workbook with xlwings
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)
        sheet_xlwings = wb_xlwings.sheets['Telephony']

        # Iterate over the rows and process based on conditions
        for row_idx in range(2, sheet_xlwings.range('A1').end('down').row + 1):  # Dynamically find last row

            # Define cells to update
            cells_to_update = ['E', 'F', 'G',]
            for col in cells_to_update:
                cell_address = f'{col}{row_idx}'
                # Get the result and replace formula with result
                cell_result = sheet_xlwings.range(cell_address).value
                sheet_xlwings.range(cell_address).value = cell_result

        # Save and close the updated workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        st.info("Replaced data with original value results. Temporary save!")
        
        #====================================================================================================

        #============================droppingUnwantedRows====================================================

        # Open the workbook with xlwings
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)
        sheet_xlwings = wb_xlwings.sheets['Telephony']  # Adjust sheet name if necessary


        # Drop the columns "Concat", "Device OS", "Asset Integrated in QRadar", "Logs Sending Status"
        columns_to_delete = ['H', 'H']
        for col in reversed(columns_to_delete):
            sheet_xlwings.range(f'{col}:{col}').delete()

        # Save and close the updated workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        st.info("Dropped unwanted data from Telephony worksheet. Temporary save!")

        #====================================================================================================

        #==============================copytelephonytoassetlist====================================================

        # Open the workbook with xlwings
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)

        # Load both worksheets
        sheet_asset_list = wb_xlwings.sheets['Asset List']
        sheet_telephony = wb_xlwings.sheets['Telephony']

        # Find the last row in the Asset List sheet (to know where to start pasting)
        last_row_asset_list = sheet_asset_list.range('A1').end('down').row + 1  # Get the first empty row

        # Find the range of data to copy from the Telephony sheet
        last_row_telephony = sheet_telephony.range('A1').end('down').row  # Get the last row in Telephony

        # Define the range of the data to copy from the Telephony sheet
        range_to_copy = sheet_telephony.range(f'A2:AB{last_row_telephony}')  # Adjust if necessary

        # Copy the range from Telephony and paste it in the Asset List sheet
        destination_range = sheet_asset_list.range(f'A{last_row_asset_list}')

        # Paste the copied range into the destination
        range_to_copy.api.Copy(destination_range.api)

        # Save and close the updated workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        st.info("Data from Telephony worksheet cloned to Asset List worksheet. Temporary save!")

        #====================================================================================================

        #=====================================hideunwantedsheets============================================
        # Load the workbook
        app = xw.App(visible=False)
        wb_xlwings = app.books.open(temp_qformula_path)

        # List of sheet names to hide
        sheets_to_hide = ['Telephony', 'Log Source']

        # Hide the specified sheets
        for sheet_name in sheets_to_hide:
            if sheet_name in [sheet.name for sheet in wb_xlwings.sheets]:
                sheet = wb_xlwings.sheets[sheet_name]
                sheet.api.Visible = 0  # 0 means hidden, 1 means visible, 2 means very hidden

        # Save the workbook
        wb_xlwings.save()
        wb_xlwings.close()
        app.quit()

        #============================================================================================

        #============================downloadintegrationreport=======================================

        # Create the new file name with the current date
        current_date = datetime.now().strftime('%d-%m-%Y')
        intg_file_name = f'IntegrationReport_{current_date}.xlsx'

        # Download temps
        st.success("Data process successful & ready for download.")
        with open(temp_qformula_path, 'rb') as f:
            st.download_button('Download processed IntegrationReport', f, file_name=intg_file_name)

        #==========================================opsEnd================================================

    else:
        st.error("Invalid. upload required files.")

    
    st.write("")
    st.write("")
    st.write("")

    # Back button
    if st.button("Back to Index"):
        st.session_state.page = 'index'

    # Label
    st.markdown(
        """
        <style>
        .footer {
            position: fixed;
            bottom: 0;
            left: 0;
            width: 100%;
            text-align: center;
            padding: 10px;
            background-color: #0E1117;
            color: #fff;
            font-size: 14px;
        }
        </style>
        <div class="footer">
            ᚱᛟᛟᛏ@ᛞᛖᛖ:~#
        </div>
        """,
        unsafe_allow_html=True
    )
